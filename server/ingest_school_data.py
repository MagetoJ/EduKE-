import os
import sys
import asyncio
from datetime import datetime
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
from sqlalchemy.orm import sessionmaker
from sqlalchemy import select, insert
from dotenv import load_dotenv

# Try importing openpyxl
try:
    import openpyxl
except ImportError:
    print("❌ ERROR: 'openpyxl' package is missing. Run: pip install openpyxl")
    sys.exit(1)

# Ensure local imports work cleanly
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from models import Base, School, User, Student, school_users, UserRole
from models_roles import ClassTeacherAssignment

try:
    from auth import get_password_hash
except ImportError:
    from passlib.context import CryptContext
    pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
    def get_password_hash(password: str) -> str:
        return pwd_context.hash(password)

load_dotenv()

DATABASE_URL = os.getenv("DATABASE_URL", "").strip()
if not DATABASE_URL:
    print("❌ ERROR: DATABASE_URL is not set in your .env file.")
    sys.exit(1)

if DATABASE_URL.startswith("postgresql://"):
    DATABASE_URL = DATABASE_URL.replace("postgresql://", "postgresql+asyncpg://", 1)

engine = create_async_engine(DATABASE_URL, echo=False)
AsyncSessionLocal = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

def parse_excel_sheet(sheet):
    """Helper to convert an openpyxl worksheet into a list of clean dictionaries, skipping header header rows"""
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    
    # Locate the true header row (the one containing column headers like Full Name or UPI ID)
    header_idx = 0
    for idx, row in enumerate(rows):
        if any(keyword in [str(cell).strip() for cell in row if cell] for keyword in ["Teacher Name", "Full Name", "UPI ID", "Learner Full Name"]):
            header_idx = idx
            break
            
    headers = [str(cell).strip() if cell else f"col_{i}" for i, cell in enumerate(rows[header_idx])]
    
    data_rows = []
    for row in rows[header_idx + 1:]:
        if not any(row): # Skip entirely blank rows
            continue
        row_dict = {}
        for i, cell in enumerate(row):
            if i < len(headers):
                val = str(cell).strip() if cell is not None else ""
                row_dict[headers[i]] = val
        data_rows.append(row_dict)
    return data_rows

async def ingest_all(school_name: str, admin_email: str, excel_filename: str):
    if not os.path.exists(excel_filename):
        print(f"❌ ERROR: Master workbook '{excel_filename}' not found in the server directory.")
        return

    print(f"📖 Opening Master Excel Workbook: {excel_filename}...")
    wb = openpyxl.load_workbook(excel_filename, data_only=True)

    async with AsyncSessionLocal() as db:
        slug = school_name.lower().replace(" ", "-")
        print(f"Connecting to database for school slug: '{slug}'...")
        
        # 1. Get or Create the School Tenant
        existing_school = await db.execute(select(School).where(School.slug == slug))
        school = existing_school.scalar_one_or_none()
        
        if school:
            print(f"ℹ️ Info: School '{school_name}' already exists (ID: {school.id}). Appending missing rows...")
        else:
            school = School(
                name=school_name,
                slug=slug,
                email=admin_email,
                status="active",
                is_special_needs=False,
                disability_category="none"
            )
            db.add(school)
            await db.flush() 
            print(f"✅ Created fresh school profile with ID: {school.id}")

        # 2. Process Teachers Sheet
        print("\nProcessing Teachers Sheet...")
        teacher_sheet = next((wb[s] for s in wb.sheetnames if "teacher" in s.lower()), None)
        if teacher_sheet is None:
            print("❌ ERROR: Could not find a sheet containing 'Teachers' inside the workbook.")
        else:
            teachers_data = parse_excel_sheet(teacher_sheet)
            teacher_count = 0
            for row in teachers_data:
                name = row.get("Teacher Name")
                email = row.get("Email Address")
                if not name or not email:
                    continue
                
                existing = await db.execute(select(User).where(User.email == email))
                if existing.scalar_one_or_none():
                    continue

                designation = row.get("Designation/Rank", "").lower()
                role = UserRole.TEACHER
                if "headteacher" in designation:
                    role = UserRole.ADMIN
                elif "class teacher" in designation:
                    role = UserRole.CLASS_TEACHER

                user = User(
                    username=email,
                    email=email,
                    full_name=name,
                    hashed_password=get_password_hash("Temporary123!"),
                    is_active=True
                )
                db.add(user)
                await db.flush()

                await db.execute(
                    insert(school_users).values(
                        school_id=school.id,
                        user_id=user.id,
                        role=role,
                        is_active=True
                    )
                )

                if role == UserRole.CLASS_TEACHER and row.get("Primary Level Tag"):
                    db.add(ClassTeacherAssignment(
                        school_id=school.id,
                        teacher_id=user.id,
                        grade_level=row["Primary Level Tag"],
                        stream_section="East"
                    ))
                teacher_count += 1
            print(f"✅ Successfully added {teacher_count} new teacher portal accounts.")

        # 3. Process Support Staff Sheet
        print("\nProcessing Support Staff Sheet...")
        staff_sheet = next((wb[s] for s in wb.sheetnames if "staff" in s.lower() or "support" in s.lower()), None)
        if staff_sheet is None:
            print("❌ ERROR: Could not find a sheet containing 'Support Staff' inside the workbook.")
        else:
            staff_data = parse_excel_sheet(staff_sheet)
            staff_count = 0
            for row in staff_data:
                name = row.get("Full Name")
                if not name:
                    continue

                safe_username = name.lower().replace(" ", "") + row.get("Staff ID", "nts")
                email = f"{safe_username}@school.ac.ke"
                
                existing = await db.execute(select(User).where(User.email == email))
                if existing.scalar_one_or_none():
                    continue

                user = User(
                    username=email,
                    email=email,
                    full_name=name,
                    hashed_password=get_password_hash("Temporary123!"),
                    is_active=True
                )
                db.add(user)
                await db.flush()

                await db.execute(
                    insert(school_users).values(
                        school_id=school.id,
                        user_id=user.id,
                        role=UserRole.STAFF,
                        is_active=True
                    )
                )
                staff_count += 1
            print(f"✅ Successfully added {staff_count} new support staff accounts.")

        # 4. Process Student Registry Sheet
        print("\nProcessing Student Registry Sheet...")
        student_sheet = next((wb[s] for s in wb.sheetnames if "student" in s.lower() or "learner" in s.lower()), None)
        if student_sheet is None:
            print("❌ ERROR: Could not find a sheet containing 'Students' inside the workbook.")
        else:
            students_data = parse_excel_sheet(student_sheet)
            student_count = 0
            for row in students_data:
                name = row.get("Learner Full Name")
                if not name:
                    continue

                name_parts = name.split(" ")
                first_name = name_parts[0]
                last_name = " ".join(name_parts[1:]) if len(name_parts) > 1 else "Learner"

                try:
                    balance = float(row.get("Fee Balance (KES)", "0").replace(",", ""))
                except ValueError:
                    balance = 0.0

                student = Student(
                    school_id=school.id,
                    first_name=first_name,
                    last_name=last_name,
                    grade=row.get("CBE Level", "Grade 7"),
                    stream_section=row.get("Stream/Class", "A"),
                    admission_number=row.get("UPI ID") or row.get("Nemis Number"),
                    status="active",
                    current_balance=balance
                )
                db.add(student)
                student_count += 1
            print(f"✅ Successfully added {student_count} new student profiles.")

        await db.commit()
        print(f"\n🎉 SUCCESS: Master Workbook data has been fully populated into '{school_name}'!")

if __name__ == "__main__":
    asyncio.run(ingest_all(
        school_name="Machakos Day Academy",
        admin_email="principal@machakosday.ac.ke",
        excel_filename="Comprehensive_Kenyan_CBE_Day_School_Data.xlsx"
    ))